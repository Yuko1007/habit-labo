from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

continue_data = [
    ('山岡 千鶴', 'RB大宮アルディージャWOMEN U15'),
    ('中来田 空奈', 'FC．TIARA GIRLS'),
    ('松本 明花里', '北海道リラ・コンサドーレ'),
    ('飯田 莉桜', '札幌市立羊丘中学校'),
    ('村松 美音', 'RENUOVENS OGASA FOOTBALL CLUB'),
    ('西村 祐蒼', '日テレ・東京ヴェルディメニーナ'),
    ('森井 咲', '三菱重工浦和レッズレディースジュニアユース'),
    ('松本 芽生', 'ちふれASエルフェン埼玉マリU-15'),
    ('加藤 真央', '朝日インテック・ラブリッジ名古屋スターチス'),
    ('古川 心都', 'セレッソ大阪ヤンマーガールズU-15'),
    ('小池 結芽', 'INAC神戸テゾーロ'),
    ('山根 杏南', 'FC.REVO山口 U-15'),
    ('鬼塚 樹南', 'サンフレッチェ福山レジーナジュニアユース'),
    ('神野 咲子', '愛媛FCレディースMIKAN//JFAアカデミー今治'),
]

dropped_data = [
    ('清水 あいこ', '北海道リラ・コンサドーレ'),
    ('田中 燈里', '三菱重工浦和レッズレディースジュニアユース'),
    ('鷲谷 心晴', '三菱重工浦和レッズレディースジュニアユース'),
    ('森川 陽茉莉', '北陸大学フィオリーレ'),
    ('西澤 琉愛', 'ジェフ千葉レディースU-15'),
    ('横木 あづ', 'アルビレックス新潟レディースU-15'),
    ('小野 こまち', '丸亀ELF女子F.C'),
    ('佐藤 愛夏', 'NJSS'),
    ('鈴木 日莉', 'グロースフットボールクラブ'),
    ('窪田 心優', '横須賀シーガルズBES'),
    ('北山 実生', '八女学院フットボールクラブ'),
    ('岩佐 梨乃花', 'セレッソ大阪ヤンマーガールズU-15'),
    ('梅本 雫', 'INAC神戸テゾーロ'),
    ('行徳 乃愛', '神村学園中'),
    ('伊東 空', 'ジェフ千葉レディースU-15'),
    ('森 菜槻', 'SATOなでしこ U-15'),
    ('満屋 桃杏', '日テレ・東京ヴェルディメニーナ'),
    ('野口 凛璃愛', 'SATOなでしこ U-15'),
]

new_data = [
    ('伊東 陽葵', '北海道リラ・コンサドーレ'),
    ('安田 歩夢', '北海道リラ・コンサドーレ'),
    ('松坂 綾音', '北海道リラ・コンサドーレ'),
    ('板元 波凪', 'マイナビ仙台レディースジュニアユース'),
    ('武士俣 咲', 'マイナビ仙台レディースジュニアユース'),
    ('小竹 泉実', 'マイナビ仙台レディースジュニアユース'),
    ('阿部 心桜', 'マイナビ仙台レディースジュニアユース'),
    ('鴇崎 妃南', '日テレ・東京ヴェルディメニーナ'),
    ('廣末 結麻', '日テレ・東京ヴェルディメニーナ'),
    ('大村 遥夏', '三菱重工浦和レッズレディースジュニアユース'),
    ('豊田 愛実', 'ジェフユナイテッド市原・千葉レディースU-15'),
    ('江川 千聖', 'INAC白岡SCレディース'),
    ('川崎 理桜', '小美玉フットボールアカデミー'),
    ('佐藤 紗奈', '日本航空高等学校附属中学校'),
    ('金丸 乃愛', '富山新庄クラブU-15'),
    ('小山 莉花', 'アルビレックス新潟レディースU-15'),
    ('丸山 まどか', 'アルビレックス新潟レディースU-15'),
    ('樋口 明莉咲', 'アルビレックス新潟レディースU-15'),
    ('玉木 蘭', 'FC福井マリーナ/JFAアカデミー堺'),
    ('名知 柚音', '岐阜西SCサッカークラブ'),
    ('小島 琉々巴', '朝日インテック・ラブリッジ名古屋スターチス'),
    ('稲垣 ひまり', 'FC．フェルボール．MIMOSA'),
    ('柴田 小湖', '尾張フットボールクラブレディース/JFAアカデミー堺'),
    ('木村 柚心', '尾張フットボールクラブレディース'),
    ('落合 裕美', '藤枝順心サッカークラブジュニアユース'),
    ('井藤 未来', 'セレッソ大阪ヤンマーガールズU-15'),
    ('高木 彩乃', 'セレッソ大阪ヤンマーガールズU-15'),
    ('藤井 菫', 'INAC神戸テゾーロ'),
    ('塩野 史季', 'INAC神戸テゾーロ'),
    ('末次 さくら', 'ディオッサ出雲FCジュニアユース'),
    ('蒔田 心都', 'AICグラーロス広島レディース'),
    ('石川 舞凛', 'FC今治レディースNEXT'),
    ('タン エリカ', 'F.C.コーマレディースアザレア'),
    ('秋岡 芽依', 'FC STORY Tokushima メニーナ'),
    ('尾崎 蘭', 'プルミエール徳島サッカークラブ'),
    ('宮田 芽依', '口石フットボールクラブ'),
    ('首藤 舞空', '大分トリニータ レディース'),
    ('奥本 琉穂', '八女学院女子フットボールクラブ'),
    ('辻元 真歩', '神村学園中等部女子サッカー部'),
    ('松岡 果歩', '神村学園中等部女子サッカー部'),
    ('平良 理葵', 'FC琉球さくらale U-15'),
    ('前川 彩音', '太陽スポーツクラブ熊本玉名U-15'),
]

wb = Workbook()
sheet = wb.active
sheet.title = 'トレセンメンバー変動'

title_font = Font(bold=True, size=14, color='FFFFFF')
title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=10)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
continue_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
dropped_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
new_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

sheet.merge_cells('A1:C1')
title_cell = sheet['A1']
title_cell.value = 'U-14ナショナルトレセン メンバー変動（U-13比較）'
title_cell.font = title_font
title_cell.fill = title_fill
sheet.row_dimensions[1].height = 25

row = 5
sheet['A' + str(row)].value = '✓ 継続者（U-13→U-14）'
sheet['A' + str(row)].font = Font(bold=True, size=12, color='FFFFFF')
sheet['A' + str(row)].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
sheet.merge_cells('A' + str(row) + ':C' + str(row))

row += 1
sheet['A' + str(row)].value = '選手名'
sheet['B' + str(row)].value = 'チーム名'
sheet['C' + str(row)].value = '区分'
for col in ['A', 'B', 'C']:
    sheet[col + str(row)].font = header_font
    sheet[col + str(row)].fill = header_fill

for name, team in continue_data:
    row += 1
    sheet['A' + str(row)].value = name
    sheet['B' + str(row)].value = team
    sheet['C' + str(row)].value = 'GK' if name in ['山岡 千鶴', '中来田 空奈'] else 'FP'
    for col in ['A', 'B', 'C']:
        sheet[col + str(row)].fill = continue_fill

row += 2
sheet['A' + str(row)].value = '✗ 外れた人（U-13にいたがU-14にいない）'
sheet['A' + str(row)].font = Font(bold=True, size=12, color='FFFFFF')
sheet['A' + str(row)].fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
sheet.merge_cells('A' + str(row) + ':C' + str(row))

row += 1
sheet['A' + str(row)].value = '選手名'
sheet['B' + str(row)].value = 'チーム名'
sheet['C' + str(row)].value = '区分'
for col in ['A', 'B', 'C']:
    sheet[col + str(row)].font = header_font
    sheet[col + str(row)].fill = header_fill

for name, team in dropped_data:
    row += 1
    sheet['A' + str(row)].value = name
    sheet['B' + str(row)].value = team
    sheet['C' + str(row)].value = 'GK' if name in ['清水 あいこ', '田中 燈里'] else 'FP'
    for col in ['A', 'B', 'C']:
        sheet[col + str(row)].fill = dropped_fill

row += 2
sheet['A' + str(row)].value = '◆ 新規入選（U-14新メンバー）'
sheet['A' + str(row)].font = Font(bold=True, size=12, color='FFFFFF')
sheet['A' + str(row)].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
sheet.merge_cells('A' + str(row) + ':C' + str(row))

row += 1
sheet['A' + str(row)].value = '選手名'
sheet['B' + str(row)].value = 'チーム名'
sheet['C' + str(row)].value = '区分'
for col in ['A', 'B', 'C']:
    sheet[col + str(row)].font = header_font
    sheet[col + str(row)].fill = header_fill

new_gk = ['伊東 陽葵', '板元 波凪', '金丸 乃愛', '名知 柚音', '石川 舞凛', '宮田 芽依']
for name, team in new_data:
    row += 1
    sheet['A' + str(row)].value = name
    sheet['B' + str(row)].value = team
    sheet['C' + str(row)].value = 'GK' if name in new_gk else 'FP'
    for col in ['A', 'B', 'C']:
        sheet[col + str(row)].fill = new_fill

sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 45
sheet.column_dimensions['C'].width = 8

summary = wb.create_sheet('サマリー')
summary['A1'].value = 'U-14ナショナルトレセン メンバー比較サマリー'
summary['A1'].font = title_font
summary['A1'].fill = title_fill
summary.merge_cells('A1:B1')

row = 3
summary['A' + str(row)].value = '項目'
summary['B' + str(row)].value = '人数'
summary['A' + str(row)].font = header_font
summary['A' + str(row)].fill = header_fill
summary['B' + str(row)].font = header_font
summary['B' + str(row)].fill = header_fill

data = [
    ('U-13→U-14 継続', len(continue_data)),
    ('新規入選', len(new_data)),
    ('外れた人', len(dropped_data)),
    ('U-14メンバー総数', 56)
]

for idx, (label, count) in enumerate(data, start=4):
    summary['A' + str(row + idx)].value = label
    summary['B' + str(row + idx)].value = count

summary.column_dimensions['A'].width = 25
summary.column_dimensions['B'].width = 12

# Part 2分析シートを追加
part2_data = {
    'GK': [('山岡 千鶴', 'RB大宮アルディージャWOMEN U15'),
           ('中来田 空奈', 'FC.TIARA GIRLS')],
    'FP': [('武士俣 咲', 'マイナビ仙台レディースジュニアユース'),
           ('鷲谷 心晴', '三菱重工浦和レッズレディースジュニアユース'),
           ('森川 陽茉莉', '北陸大学フィオリーレ'),
           ('西澤 琉愛', 'ジェフユナイテッド市原・千葉レディースU-15'),
           ('古川 心都', 'セレッソ大阪ヤンマーガールズU-15'),
           ('伊藤 葵海', '日テレ・東京ヴェルディメニーナ'),
           ('小野 こまち', '丸亀ELF女子F.C／JFAアカデミー今治'),
           ('岩佐 梨乃花', 'セレッソ大阪ヤンマーガールズU-15'),
           ('小池 結芽', 'INAC神戸テゾーロ'),
           ('梅本 雫', 'INAC神戸テゾーロ'),
           ('行德 乃愛', '神村学園中'),
           ('奥本 琉穂', '八女学院女子フットボールクラブ'),
           ('後藤 月乃', '岐阜西SCサッカークラブ／JFAアカデミー堺'),
           ('松坂 綾音', '北海道リラ・コンサドーレ'),
           ('首藤 舞空', '大分トリニータ レディース'),
           ('西村 祐蒼', '日テレ・東京ヴェルディメニーナ'),
           ('清水 杏奈', 'サンフレッチェ広島レジーナ ジュニアユース'),
           ('山根 杏南', 'FC.REVO山口 U-15')]
}

u14_members = set([
    '伊東 陽葵', '松本 明花里', '安田 歩夢', '松坂 綾音', '飯田 莉桜',
    '板元 波凪', '村松 美音', '武士俣 咲', '小竹 泉実', '阿部 心桜',
    '山岡 千鶴', '西村 祐蒼', '鴇崎 妃南', '廣末 結麻', '大村 遥夏',
    '森井 咲', '豊田 愛実', '江川 千聖', '松本 芽生', '川崎 理桜',
    '佐藤 紗奈', '金丸 乃愛', '小山 莉花', '丸山 まどか', '樋口 明莉咲',
    '玉木 蘭', '名知 柚音', '小島 琉々巴', '加藤 真央', '稲垣 ひまり',
    '柴田 小湖', '木村 柚心', '落合 裕美', '中来田 空奈', '古川 心都',
    '井藤 未来', '高木 彩乃', '小池 結芽', '藤井 菫', '塩野 史季',
    '山根 杏南', '鬼塚 樹南', '末次 さくら', '蒔田 心都', '石川 舞凛',
    'タン エリカ', '秋岡 芽依', '尾崎 蘭', '神野 咲子', '首藤 舞空',
    '奥本 琉穂', '辻元 真歩', '松岡 果歩', '平良 理葵', '前川 彩音'
])

# Part 2分析
part2_sheet = wb.create_sheet('U-13 Part2分析')
part2_sheet['A1'].value = 'U-13 Part2とU-14の重複分析'
part2_sheet['A1'].font = title_font
part2_sheet['A1'].fill = title_fill
part2_sheet.merge_cells('A1:C1')

row = 3
part2_sheet['A' + str(row)].value = '✓ 継続者（U-13 Part2→U-14）'
part2_sheet['A' + str(row)].font = Font(bold=True, size=12, color='FFFFFF')
part2_sheet['A' + str(row)].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
part2_sheet.merge_cells('A' + str(row) + ':C' + str(row))

row += 1
part2_sheet['A' + str(row)].value = '選手名'
part2_sheet['B' + str(row)].value = 'チーム名'
part2_sheet['C' + str(row)].value = '区分'
for col in ['A', 'B', 'C']:
    part2_sheet[col + str(row)].font = header_font
    part2_sheet[col + str(row)].fill = header_fill

for pos, players in part2_data.items():
    for name, team in players:
        if name in u14_members:
            row += 1
            part2_sheet['A' + str(row)].value = name
            part2_sheet['B' + str(row)].value = team
            part2_sheet['C' + str(row)].value = pos
            for col in ['A', 'B', 'C']:
                part2_sheet[col + str(row)].fill = continue_fill

row += 2
part2_sheet['A' + str(row)].value = '✗ 外れた人（U-13 Part2にいたがU-14にいない）'
part2_sheet['A' + str(row)].font = Font(bold=True, size=12, color='FFFFFF')
part2_sheet['A' + str(row)].fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
part2_sheet.merge_cells('A' + str(row) + ':C' + str(row))

row += 1
part2_sheet['A' + str(row)].value = '選手名'
part2_sheet['B' + str(row)].value = 'チーム名'
part2_sheet['C' + str(row)].value = '区分'
for col in ['A', 'B', 'C']:
    part2_sheet[col + str(row)].font = header_font
    part2_sheet[col + str(row)].fill = header_fill

for pos, players in part2_data.items():
    for name, team in players:
        if name not in u14_members:
            row += 1
            part2_sheet['A' + str(row)].value = name
            part2_sheet['B' + str(row)].value = team
            part2_sheet['C' + str(row)].value = pos
            for col in ['A', 'B', 'C']:
                part2_sheet[col + str(row)].fill = dropped_fill

part2_sheet.column_dimensions['A'].width = 15
part2_sheet.column_dimensions['B'].width = 45
part2_sheet.column_dimensions['C'].width = 8

wb.save(r'C:\Users\somey\Documents\private\U14_full_analysis.xlsx')
print('Complete')
